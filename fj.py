import tkinter as fj 
import openpyxl as miss
from tkinter import messagebox, ttk

def display():
    wb = miss.load_workbook("Miss CPA.xlsx")
    sheet = wb.active
    for row in table.get_children():
        table.delete(row)
        
    for row in sheet.iter_rows(min_row = 2, values_only = True):
        table.insert("", fj.END, values = row)
def input_validation():
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birth = birth_entry.get()
    
    if not first or not middle or not last or not birth:
        messagebox.showerror("Error", "All fields are required!")
        return False
    
    if not birth.isdigit():
        messagebox.showerror("Error", "Birth year must be in a number form!")
        return False
    return True

def saving():
    if not input_validation():
        return
    
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birth = int(birth_entry.get())
    age = 2026 - birth
    
    wb = miss.load_workbook("Miss CPA.xlsx")
    sheet = wb.active
    
    new_id = sheet.max_row
    
    sheet.append([new_id , last, first, middle, birth, age])
    wb.save("Miss CPA.xlsx")
    
    messagebox.showinfo("Success", "Record added successfully!")
    display()
def auto_populated(event):
    selected = table.focus()
    values = table.item(selected, "values")
    
    if values:
        fname_entry.delete(0, fj.END)
        mname_entry.delete(0, fj.END)
        lname_entry.delete(0, fj.END)
        birth_entry.delete(0, fj.END)
        
        fname_entry.insert(0, values[2])
        mname_entry.insert(0, values[3])
        lname_entry.insert(0, values[1]) 
        birth_entry.insert(0, values[4])      
def update():
    selected = table.focus()
    
    if not selected:
        messagebox.showerror("Error", "Select a record first.")
        return 
    if not input_validation():
        return
    
    values = table.item(selected, "values")
    record_id = values[0]
    
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birth = int(birth_entry.get())
    
    age = 2026 - birth
    
    wb = miss.load_workbook("Miss CPA.xlsx")
    sheet = wb.active
    
    for row in sheet.iter_rows(min_row = 2):
        
        if str(row[0].value)==str(record_id):
            row[1].value = last
            row[2].value = first
            row[3].value = middle
            row[4].value = birth
            row[5].value = age
            
    wb.save("Miss CPA.xlsx")
    messagebox.showinfo("Success", "Record updated successfully!")
    display()
    
def delete():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first")
        return

    values = table.item(selected, "values")
    record_id = values[0]

    confirm = messagebox.askyesnocancel("Confirm","Are you sure you want to delete this record?")

    if not confirm:
        return

    wb = miss.load_workbook("Miss CPA.xlsx")
    sheet = wb.active

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if str(row[0].value) == str(record_id):
            sheet.delete_rows(i)
            break

    wb.save("Miss CPA.xlsx")
    messagebox.showinfo("Success", "Record deleted successfully!")
    display()
    
window = fj.Tk()
window.title("Age Calculator")
window.configure(bg="light pink")

title = fj.Label (window, text= "Profile Builder", font =("Times New Roman", 14, "bold"), bg = "light pink")
title.grid(row = 0, column = 0, columnspan = 6)

gen_frame = fj.Frame(window, bg = "light pink", bd = 2, relief = "groove")
gen_frame.grid(row = 1, column = 0, columnspan = 6, padx = 10, pady = 10)

fname_entry = fj.Entry(gen_frame, font = ("Poppins", 12))
fname_entry.grid(row = 2, column = 1, columnspan = 2, padx = (10,0), pady = (10,0))

fname_label = fj.Label(gen_frame, text = "First name", font = ("Poppins", 10, "bold"), bg = "Light Pink")
fname_label.grid(row = 3, column = 1, columnspan = 2)

mname_entry = fj.Entry(gen_frame, font = ("Poppins", 12))
mname_entry.grid(row = 2, column = 3, columnspan = 2, padx = (10,0), pady = (10,0))

mname_label = fj.Label(gen_frame, text = "Middle name", font = ("Poppins", 10, "bold"), bg = "Light Pink")
mname_label.grid(row = 3, column = 3, columnspan = 2)

lname_entry = fj.Entry(gen_frame, font = ("Poppins", 12))
lname_entry.grid(row = 2, column = 5, columnspan = 2, padx = (10,0), pady = (10,0))

lname_label = fj.Label(gen_frame, text = "Last name", font = ("Poppins", 10, "bold"), bg = "Light Pink")
lname_label.grid(row = 3, column = 5, columnspan = 2)

birth_entry = fj.Entry(gen_frame, font = ("Poppins", 12))
birth_entry.grid(row = 4, column = 1, columnspan = 2, padx = (10,0), pady = (10,0))

birth_year_label = fj.Label(gen_frame, text = "Birth Year", font = ("Poppins", 10, "bold"), bg = "Light Pink")
birth_year_label.grid(row = 5, column = 1, columnspan = 2)

update_btn =  fj.Button(window, text = "Update", command = update)
update_btn.grid(row = 6, column = 2)

btn = fj.Button(window, text = "Submit", command = saving, font = ("Poppins", 12, "bold"), bg = "lightblue")
btn.grid(row = 6, column = 0, columnspan = 6, pady = (10,20))

delete_btn = fj.Button(window, text = "Delete", command = delete)
delete_btn.grid(row = 6, column = 3)

table = ttk.Treeview(window, columns = ("ID", "Last", "First", "Middle", "Birth", "Age"), show = "headings")
for col in ("ID", "Last", "First", "Middle", "Birth", "Age"):
    table.heading(col, text = col)
table.grid(row = 7, column = 0, columnspan = 4)

table.bind("<<TreeviewSelect>>", auto_populated)


wb = miss.Workbook()
sheet = wb.active

sheet.title = "Sheet1"

sheet.append(["ID", "Last", "First", "Middle", "Birth", "Age"])

wb.save("Miss CPA.xlsx")

print("Excel file created successfully!")

display()
window.mainloop()
    