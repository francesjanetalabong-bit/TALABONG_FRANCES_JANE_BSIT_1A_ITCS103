from openpyxl import Workbook as fj
from datetime import datetime as dt

wb = fj()
ws = wb.active
ws.title = "Favorite People"

ws['A1'] = "ID"
ws['B1'] = "First Name"
ws['C1'] = "Middle Name"
ws['D1'] = "Last Name"
ws['E1'] = "Birth Date"
ws['F1'] = "Age"

for i in range(1, 4):
    print(f"\nEnter information for person #{i}:")

    first_name = input("First Name: ")
    middle_name = input("Middle Name: ")
    last_name = input("Last Name: ")

    while True:
        try:
            birth_date_input = input("Birth Date (YYYY-MM-DD) (Ex. 2000-01-01): ")
            birth_date = dt.strptime(birth_date_input, "%Y-%m-%d")
            break

        except ValueError:
            print("Invalid input. Please enter a valid birth date in YYYY-MM-DD format.")

    today = dt.today()

    age = today.year - birth_date.year

    if (today.month, today.day) < (birth_date.month, birth_date.day):
        age -= 1

    person_id = i

    ws.cell(row=i + 1, column=1, value=person_id)
    ws.cell(row=i + 1, column=2, value=first_name)
    ws.cell(row=i + 1, column=3, value=middle_name)
    ws.cell(row=i + 1, column=4, value=last_name)
    ws.cell(row=i + 1, column=5, value=birth_date_input)
    ws.cell(row=i + 1, column=6, value=age)

file_name = "special_people.xlsx"
wb.save(file_name)

print(f"\nAll data saved successfully in '{file_name}'!\n")

print("Saved Records:")
for row in ws.iter_rows(values_only=True):
    print(row)